# -*- coding: utf-8 -*-

# Form implementation generated from reading ui file 'Springer XLSX Beta.ui'
#
# Created by: PyQt5 UI code generator 5.13.2
#
# WARNING! All changes made in this file will be lost!
import traceback
from tkinter import Tk, filedialog

import requests
from PyQt5 import QtCore, QtGui, QtWidgets
import openpyxl
from pathlib import Path

from PyQt5.QtCore import QObject, pyqtSignal, QRunnable, pyqtSlot, QThreadPool
from PyQt5.QtGui import QIcon
from openpyxl import Workbook
done_html = 0
global_col = []
wb = Workbook()
ws =  wb.active
ws.title = "Links"
current_link = 0
class WorkerSignals(QObject):
    '''
    Defines the signals available from a running worker thread.

    Supported signals are:

    finished
        No data

    error
        `tuple` (exctype, value, traceback.format_exc() )

    result
        `object` data returned from processing, anything

    progress
        `int` indicating % progress

    '''
    finished = pyqtSignal()
    error = pyqtSignal(tuple)
    result = pyqtSignal(object)
    progress = pyqtSignal(int)


class Worker(QRunnable):
    '''
    Worker thread

    Inherits from QRunnable to handler worker thread setup, signals and wrap-up.

    :param callback: The function callback to run on this worker thread. Supplied args and
                     kwargs will be passed through to the runner.
    :type callback: function
    :param args: Arguments to pass to the callback function
    :param kwargs: Keywords to pass to the callback function

    '''

    def __init__(self, fn, *args, **kwargs):
        super(Worker, self).__init__()

        # Store constructor arguments (re-used for processing)
        self.fn = fn
        self.args = args
        self.kwargs = kwargs
        self.signals = WorkerSignals()

        # Add the callback to our kwargs
        self.kwargs['progress_callback'] = self.signals.progress #check progress_callbank

    @pyqtSlot()
    def run(self):
        '''
        Initialise the runner function with passed args, kwargs.
        '''
        # Retrieve args/kwargs here; and fire processing using them
        try:
            result = self.fn(*self.args, **self.kwargs)
        except:
            traceback.print_exc()
            exctype, value = sys.exc_info()[:2]
            self.signals.error.emit((exctype, value, traceback.format_exc()))
        else:
            self.signals.result.emit(result)  # Return the result of the processing
        finally:
            self.signals.finished.emit()  # Done

class Ui_Dialog(object):
    def setupUi(self, Dialog):
        Dialog.setObjectName("Dialog")
        Dialog.resize(490, 139)
        self.verticalLayoutWidget = QtWidgets.QWidget(Dialog)
        self.verticalLayoutWidget.setGeometry(QtCore.QRect(9, 9, 81, 61))
        self.verticalLayoutWidget.setObjectName("verticalLayoutWidget")
        self.verticalLayout = QtWidgets.QVBoxLayout(self.verticalLayoutWidget)
        self.verticalLayout.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout.setObjectName("verticalLayout")
        self.label_2 = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.label_2.setObjectName("label_2")
        self.verticalLayout.addWidget(self.label_2)
        self.label = QtWidgets.QLabel(self.verticalLayoutWidget)
        self.label.setObjectName("label")
        self.verticalLayout.addWidget(self.label)
        self.verticalLayoutWidget_2 = QtWidgets.QWidget(Dialog)
        self.verticalLayoutWidget_2.setGeometry(QtCore.QRect(89, 9, 301, 61))
        self.verticalLayoutWidget_2.setObjectName("verticalLayoutWidget_2")
        self.verticalLayout_2 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_2)
        self.verticalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_2.setObjectName("verticalLayout_2")
        self.lineEdit = QtWidgets.QLineEdit(self.verticalLayoutWidget_2)
        self.lineEdit.setObjectName("lineEdit")
        self.verticalLayout_2.addWidget(self.lineEdit)
        self.lineEdit_2 = QtWidgets.QLineEdit(self.verticalLayoutWidget_2)
        self.lineEdit_2.setObjectName("lineEdit_2")
        self.verticalLayout_2.addWidget(self.lineEdit_2)
        self.verticalLayoutWidget_3 = QtWidgets.QWidget(Dialog)
        self.verticalLayoutWidget_3.setGeometry(QtCore.QRect(400, 10, 81, 61))
        self.verticalLayoutWidget_3.setObjectName("verticalLayoutWidget_3")
        self.verticalLayout_3 = QtWidgets.QVBoxLayout(self.verticalLayoutWidget_3)
        self.verticalLayout_3.setContentsMargins(0, 0, 0, 0)
        self.verticalLayout_3.setObjectName("verticalLayout_3")
        self.pushButton = QtWidgets.QPushButton(self.verticalLayoutWidget_3)
        self.pushButton.setObjectName("pushButton")
        self.verticalLayout_3.addWidget(self.pushButton)
        self.pushButton.clicked.connect(self.selectXLSX_Sourse)
        self.pushButton_2 = QtWidgets.QPushButton(self.verticalLayoutWidget_3)
        self.pushButton_2.setObjectName("pushButton_2")
        self.verticalLayout_3.addWidget(self.pushButton_2)
        self.pushButton_2.clicked.connect(self.selectSaveFolder)
        self.horizontalLayoutWidget = QtWidgets.QWidget(Dialog)
        self.horizontalLayoutWidget.setGeometry(QtCore.QRect(10, 70, 471, 31))
        self.horizontalLayoutWidget.setObjectName("horizontalLayoutWidget")
        self.horizontalLayout = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget)
        self.horizontalLayout.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout.setObjectName("horizontalLayout")
        self.label_3 = QtWidgets.QLabel(self.horizontalLayoutWidget)
        self.label_3.setObjectName("label_3")
        self.horizontalLayout.addWidget(self.label_3)
        self.progressBar = QtWidgets.QProgressBar(self.horizontalLayoutWidget)
        self.progressBar.setProperty("value", 0)
        self.progressBar.setObjectName("progressBar")
        self.horizontalLayout.addWidget(self.progressBar)
        self.horizontalLayoutWidget_2 = QtWidgets.QWidget(Dialog)
        self.horizontalLayoutWidget_2.setGeometry(QtCore.QRect(10, 100, 471, 31))
        self.horizontalLayoutWidget_2.setObjectName("horizontalLayoutWidget_2")
        self.horizontalLayout_2 = QtWidgets.QHBoxLayout(self.horizontalLayoutWidget_2)
        self.horizontalLayout_2.setContentsMargins(0, 0, 0, 0)
        self.horizontalLayout_2.setObjectName("horizontalLayout_2")
        self.label_4 = QtWidgets.QLabel(self.horizontalLayoutWidget_2)
        font = QtGui.QFont()
        font.setPointSize(10)
        font.setBold(True)
        font.setWeight(75)
        self.label_4.setFont(font)
        self.label_4.setObjectName("label_4")
        self.horizontalLayout_2.addWidget(self.label_4)
        self.pushButton_3 = QtWidgets.QPushButton(self.horizontalLayoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Minimum, QtWidgets.QSizePolicy.Minimum)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.pushButton_3.sizePolicy().hasHeightForWidth())
        self.pushButton_3.setSizePolicy(sizePolicy)
        self.pushButton_3.setMinimumSize(QtCore.QSize(120, 0))
        self.pushButton_3.setObjectName("pushButton_3")
        self.horizontalLayout_2.addWidget(self.pushButton_3)
        self.pushButton_3.clicked.connect(self.get_link_thread)
        self.label_5 = QtWidgets.QLabel(self.horizontalLayoutWidget_2)
        self.label_5.setAlignment(QtCore.Qt.AlignRight|QtCore.Qt.AlignTrailing|QtCore.Qt.AlignVCenter)
        self.label_5.setObjectName("label_5")
        self.horizontalLayout_2.addWidget(self.label_5)
        self.spinBox = QtWidgets.QSpinBox(self.horizontalLayoutWidget_2)
        sizePolicy = QtWidgets.QSizePolicy(QtWidgets.QSizePolicy.Maximum, QtWidgets.QSizePolicy.Fixed)
        sizePolicy.setHorizontalStretch(0)
        sizePolicy.setVerticalStretch(0)
        sizePolicy.setHeightForWidth(self.spinBox.sizePolicy().hasHeightForWidth())
        self.spinBox.setSizePolicy(sizePolicy)
        self.spinBox.setMinimum(1)
        self.spinBox.setProperty("value", 8)
        self.spinBox.setObjectName("spinBox")
        self.horizontalLayout_2.addWidget(self.spinBox)

        self.retranslateUi(Dialog)
        QtCore.QMetaObject.connectSlotsByName(Dialog)
        self.threadpool = QThreadPool()

    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        MainWindow.setWindowIcon(QIcon('lib/icon.png'))
        Dialog.setWindowTitle(_translate("Dialog", "Springer XLSX Beta"))
        self.label_2.setText(_translate("Dialog", "Source XLSX:"))
        self.label.setText(_translate("Dialog", "Output XLSX:"))
        self.pushButton.setText(_translate("Dialog", "Change ..."))
        self.pushButton_2.setText(_translate("Dialog", "Change ..."))
        self.label_3.setText(_translate("Dialog", "Progress:"))
        self.label_4.setText(_translate("Dialog", "Springer XLSX Beta"))
        self.pushButton_3.setText(_translate("Dialog", "Get Link"))
        self.label_5.setText(_translate("Dialog", "version 1.02 | Multithread:"))

    def selectXLSX_Sourse(self):
        root.filename = filedialog.askopenfilename(initialdir="/", title="Select files",
                                                    filetypes=(("XLSX Source", ".xlsx"), ("All Files", "*.*")))
        self.lineEdit.setText(root.filename)


    def selectSaveFolder(self):
        root.directory = filedialog.askdirectory()
        self.lineEdit_2.setText(root.directory)

    def get_link_thread(self):
        # Pass the function to execute
        worker = Worker(self.convert_all)  # Any other args, kwargs are passed to the run function
        worker.signals.result.connect(self.get_func_output)
        worker.signals.finished.connect(self.get_func_complete)
        worker.signals.progress.connect(self.get_proress_show)
        # Execute
        self.threadpool.start(worker)

    def convert_all(self, progress_callback):
        progress_callback.emit(0)
        xlsx_file = Path('SimData', self.lineEdit.text())
        wb_obj = openpyxl.load_workbook(xlsx_file)
        # Read the active sheet:
        sheet = wb_obj.active
        col_names = []
        for column in sheet.iter_cols(1, sheet.max_column):
            for column_p in column:
                col_names.append(column_p.value)
        global global_col
        global_col = col_names
        #print(col_names)
        for i_convert in range(0, len(col_names), self.spinBox.value()+1):
            for url_current in range(i_convert, i_convert+self.spinBox.value()+1):
                if(url_current < len(col_names)):
                    #print(url_current)
                    self.convert_thread(url_current)
                    progress_callback.emit(int ((url_current+1)*99/len(col_names)))
            global current_link
            global done_html
            while (current_link < self.spinBox.value() and done_html < len(col_names)):
                True
                #print("wait" + str(current_link))
            current_link = 0
        while (done_html < len(col_names)):
            True
        global wb
        wb.save(filename=self.lineEdit_2.text() + '/getlink.xlsx')
        progress_callback.emit(100)
        return "Done_All"

    def get_func_output(self, s):
        print(s)

    def get_func_complete(self):
        print("THREAD COMPLETE!")

    def get_proress_show(self, n):
        self.progressBar.setProperty("value", n)

    def convert_thread(self, link):
        self.link = link
        # Pass the function to execute
        worker = Worker(self.convertProgress, self.link)  # Any other args, kwargs are passed to the run function
        worker.signals.result.connect(self.print_output)
        worker.signals.finished.connect(self.thread_complete)
        worker.signals.progress.connect(self.progress_fn)
        # Execute
        self.threadpool.start(worker)

    def convertProgress(self, link, progress_callback):
        global ws
        global current_link
        global global_col
        global done_html
        progress_callback.emit(100)
        f = requests.get(global_col[link])
        html = f.text
        f_pos = html.find('<div class="cta-button-container__item">')
        f_pos = html.find('<a href="', f_pos+1)
        more = len('<a href="')
        e_pos = html.find('"', f_pos+more)
        c1 = ws.cell(row=link+1, column=1)
        c1.value = "https://link.springer.com" + html[f_pos+more:e_pos]
        #print("|"+str(link)+"|")
        # if(link == 8):
        #     print(html)
        #     print("A:"+ c1.value)
        f_pos = html.find('<div class="cta-button-container__item">', e_pos)
        f_pos = html.find('<a href="', f_pos+1)
        more = len('<a href="')
        e_pos = html.find('"', f_pos+more)
        c2 = ws.cell(row=link+1, column=2)
        c2.value = "https://link.springer.com" + html[f_pos+more:e_pos]
        current_link += 1
        done_html += 1
        return "Done"

    def print_output(self, s):
        True

    def thread_complete(self):
        True

    def progress_fn(self, n):
        True

if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    MainWindow = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(MainWindow)
    root = Tk()
    root.withdraw()
    MainWindow.show()
    sys.exit(app.exec_())